import time
import numpy as np
import imageio
from PIL import Image
from deconvolution_functions import deconv_Wiener, deconv_RL, deconv_BRL, kernal_preprocess
from TV_functions import TVL1, TVL2, TVpoisson
import os
import sys
import matplotlib.pyplot as plt
import openpyxl
INF = float("inf")
DBL_MIN = sys.float_info.min

gamma = 2.2
    
########################################################################
def read_image(file_path):
    img_in = np.asarray(Image.open(file_path))
    img_in = np.transpose(img_in, (2,0,1))

    return img_in

########################################################################
def write_image(file_path, array):
    array = np.transpose(array, (1, 2, 0))
    imageio.imwrite(file_path , array.astype(np.uint8))
    
########################################################################
def Wiener_deconv(TestImage = 'G_channel',filename = 'G.png',SNR_F=150.0):
    print ("//--------------------------------------------------------")
    print (f"start Wiener deconvolution, TEST_PAT_SIZE: {TestImage}\n")
    
    # I/O
    img_in = read_image(f'../data/{TestImage}/blur_{filename}')
    k_in = np.asarray(Image.open(f'../data/{TestImage}/kernel_{filename}'))
    
    # setting 
    to_linear = False

    # work
    t_start = time.time()
    k_in = kernal_preprocess(img_in, k_in, to_linear, gamma)
    Wiener_result = deconv_Wiener(img_in, k_in, SNR_F, to_linear, gamma)
    t_end = time.time()
    path = f'../data/{TestImage}/result/Wiener_m_SNRF{SNR_F}/{filename}'

    # 取得資料夾的路徑
    folder = os.path.dirname(path)
    
    # 如果資料夾不存在，則創建資料夾
    if not os.path.exists(folder):
        os.makedirs(folder)

    # store image
    write_image(path , Wiener_result)


########################################################################
def RL_deconv(TestImage = 'G_channel',filename = 'G.png', max_iter_RL = 25):
    print ("//--------------------------------------------------------")
    print (f"start RL-(a) deconvolution, TEST_PAT_SIZE: {TestImage}\n")
    
    # I/O
    img_in = read_image(f'../data/{TestImage}/blur_{filename}')
    k_in = np.asarray(Image.open(f'../data/{TestImage}/kernel_{filename}'))
    
    # setting 
    
    to_linear = False

    # work
    t_start = time.time()
    RL_result = deconv_RL(img_in, k_in, max_iter_RL, to_linear, gamma)
    t_end = time.time()

    path = f'../data/{TestImage}/result/RL_s_iter{max_iter_RL}/{filename}'
    # 取得資料夾的路徑
    folder = os.path.dirname(path)
    
    # 如果資料夾不存在，則創建資料夾
    if not os.path.exists(folder):
        os.makedirs(folder)

    # store image
    write_image(path , RL_result)
    
########################################################################   
def BRL_deconv(TestImage = 'G_channel',filename = 'G.png', max_iter_RL = 25, rk = 6, sigma_r = 50.0, lamb_da = 0.02):
    print ("//--------------------------------------------------------")
    print (f"start BRL-(a) deconvolution, TEST_PAT_SIZE: {TestImage}\n")
    
    # I/O
    img_in = read_image(f'../data/{TestImage}/blur_{filename}')
    k_in = np.asarray(Image.open(f'../data/{TestImage}/kernel_{filename}'))
    
    # setting
    to_linear = False
    sigma_r = sigma_r/255/255
    lamb_da = lamb_da/255

    # work
    t_start = time.time()
    BRL_result = deconv_BRL(img_in, k_in, max_iter_RL, lamb_da, sigma_r, rk, to_linear, gamma)
    t_end = time.time()

    # store image
    sigma_r = sigma_r * 255 * 255
    lamb_da = lamb_da * 255

    path = f'../data/{TestImage}/result/BRL_s_iter{max_iter_RL}_rk{rk}_si{sigma_r}_lam{lamb_da}/{filename}'
    # 取得資料夾的路徑
    folder = os.path.dirname(path)
    
    # 如果資料夾不存在，則創建資料夾
    if not os.path.exists(folder):
        os.makedirs(folder)
    

    write_image(path , BRL_result)

    

########################################################################
def TVL1_deconv(TestImage = 'G_channel',filename = 'G.png', max_iter = 1000, lamb_da = 0.01):
    print ("//--------------------------------------------------------")
    print (f"start TVL1, TEST_PAT_SIZE: {TestImage}\n")
    
    # I/O
    img_in = np.asarray(Image.open(f'../data/{TestImage}/blur_{filename}'))
    k_in = np.asarray(Image.open(f'../data/{TestImage}/kernel_{filename}'))
    
    # work
    t_start = time.time()
    TVL1_result = TVL1(img_in, k_in, max_iter, lamb_da)
    t_end = time.time()

    path = f'../data/{TestImage}/result/TVL1_m_iter{max_iter}_lam{lamb_da}/{filename}'
    # 取得資料夾的路徑
    folder = os.path.dirname(path)
    
    # 如果資料夾不存在，則創建資料夾
    if not os.path.exists(folder):
        os.makedirs(folder)
    
    # store image
    imageio.imwrite(path , TVL1_result)
    
    
########################################################################
def TVL2_deconv(TestImage = 'G_channel',filename = 'G.png', max_iter = 1000, lamb_da = 0.01):
    print ("//--------------------------------------------------------")
    print (f"start TVL2, TEST_PAT_SIZE: {TestImage}\n")
    
    # I/O
    img_in = np.asarray(Image.open(f'../data/{TestImage}/blur_{filename}'))
    k_in = np.asarray(Image.open(f'../data/{TestImage}/kernel_{filename}'))
    
    # setting
    to_linear = False

    # work
    t_start = time.time()
    TVL2_result = TVL2(img_in, k_in, max_iter, lamb_da, to_linear, gamma)
    t_end = time.time()
    path = f'../data/{TestImage}/result/TVL2_m_iter{max_iter}_lam{lamb_da}/{filename}'
    # 取得資料夾的路徑
    folder = os.path.dirname(path)
    
    # 如果資料夾不存在，則創建資料夾
    if not os.path.exists(folder):
        os.makedirs(folder)

    # store image
    imageio.imwrite(path , TVL2_result)
    

    
########################################################################
def TVpoisson_deconv(TestImage = 'G_channel',filename = 'G.png', max_iter = 1000, lamb_da = 0.01):  
    print ("//--------------------------------------------------------")
    print (f"start TVpoisson, TEST_PAT_SIZE: {TestImage}\n")
    
    # I/O
    img_in = np.asarray(Image.open(f'../data/{TestImage}/blur_{filename}'))
    k_in = np.asarray(Image.open(f'../data/{TestImage}/kernel_{filename}'))
    
    # setting
    to_linear = False

    # work
    t_start = time.time()
    TVpoisson_result = TVpoisson(img_in, k_in, max_iter, lamb_da, to_linear, gamma)
    t_end = time.time()
    path = f'../data/{TestImage}/result/TVpoisson_m_iter{max_iter}_lam{lamb_da}/{filename}'
    # 取得資料夾的路徑
    folder = os.path.dirname(path)
    
    # 如果資料夾不存在，則創建資料夾
    if not os.path.exists(folder):
        os.makedirs(folder)

    # store image
    imageio.imwrite(path , TVpoisson_result)
        
def RGB_generate(folder_path, img_list):
    image_files = [f for f in os.listdir(folder_path) if not f.startswith("RGGB") and f.endswith(".png")]
    image_files.sort(key = lambda x : int(x.split(".")[0]))
    result = np.zeros_like(read_image(os.path.join(folder_path, image_files[0])))
    count = 0
    for i in img_list:
        img = read_image(os.path.join(folder_path, image_files[i]))
        #print(os.path.join(folder_path, image_files[i]))
        count += 1
        result += img
    result = result /count
    result = result / np.max(result) *255
    result = np.roll(result,-3, axis=1)
    result = np.roll(result,-1, axis=2)
    result = result[:,8:39,8:39]
    result = result[:,::-1,::-1]
    output_file_path = os.path.join(folder_path, "RGGB.png")
    write_image(output_file_path, result)

def get_spectrum(folder_path):
    image_files = [f for f in os.listdir(folder_path) if not f.startswith("RGGB") and f.endswith(".png")]
    image_files.sort(key = lambda x : int(x.split(".")[0]))
    TR = []
    TG1 = []
    TG2 = []
    TB = []
    n = 16
    nx = 2
    ny = 2
    
    Ri_st = 8 + 8
    Rj_st = 8 + 8
    Ri_end = Ri_st + int(n//2-1)
    Rj_end = Rj_st + int(n//2-1)
    G1i_st = 0 + 8
    G1j_st = 8 + 8
    G1i_end = G1i_st + int(n//2-1)
    G1j_end = G1j_st + int(n//2-1)
    G2i_st = 8 + 8
    G2j_st = 0 + 8
    G2i_end = G2i_st + int(n//2-1)
    G2j_end = G2j_st + int(n//2-1)
    Bi_st = 0 + 8
    Bj_st = 0 + 8
    Bi_end = Bi_st + int(n//2-1)
    Bj_end = Bj_st + int(n//2-1)
    for i in range(81):
        #print(image_files[i])
        img = read_image(os.path.join(folder_path, image_files[i]))
        img = np.roll(img,-3, axis=1)
        img = np.roll(img,-2, axis=2)
        R = [np.sum(img[:,\
            Ri_st+ n*i:Ri_end + n*i,\
            Rj_st+ n*j:Rj_end + n*j]) \
            for i in range(nx) for j in range(ny)]
        G1 = [np.sum(img[:,\
            G1i_st+ n*i:G1i_end + n*i,\
            G1j_st+ n*j:G1j_end + n*j]) \
            for i in range(nx) for j in range(ny)]
        G2 = [np.sum(img[:,\
            G2i_st+ n*i:G2i_end + n*i,\
            G2j_st+ n*j:G2j_end + n*j]) \
            for i in range(nx) for j in range(ny)]
        B = [np.sum(img[:,\
            Bi_st+ n*i:Bi_end + n*i,\
            Bj_st+ n*j:Bj_end + n*j]) \
            for i in range(nx) for j in range(ny)]
        
        R = np.sum(R)
        G1 = np.sum(G1)
        G2 = np.sum(G2)
        B = np.sum(B)
        """ 
        print(R)
        print(G1)
        print(G2)
        print(B)
         """
        TR.append(R/(R+G1+G2+B))
        TG1.append(G1/(R+G1+G2+B))
        TG2.append(G2/(R+G1+G2+B))
        TB.append(B/(R+G1+G2+B))
    return TR, TG1, TG2, TB

def plot_spectrum(folder_path, wavelength, TR, TG1, TG2, TB):
    plt.figure(figsize=(10, 6))  # 設置圖像大小

    # 繪製四條曲線
    plt.plot(wavelength, TR, 'ro-', label='R')
    plt.plot(wavelength, TG1, 'go-', label='G1')
    plt.plot(wavelength, TG2, 'yo-', label='G2')
    plt.plot(wavelength, TB, 'bo-', label='B')

    # 設置坐標軸標籤
    plt.xlabel('Wavelength (nm)', fontname='Times New Roman', fontweight='bold', fontsize=24)
    plt.ylabel('Intensity (a.u.)', fontname='Times New Roman', fontweight='bold', fontsize=24)
    plt.title('Spectrum', fontname='Times New Roman', fontweight='bold', fontsize=24)
    plt.ylim(ymin = 0)
    plt.ylim(ymax = 0.5)

    # 設置刻度標籤樣式
    plt.tick_params(axis='both', which='major', labelsize=20, direction='in')

    # 添加圖例
    plt.legend(fontsize=16)

    # 保存圖像為JPG文件
    image_path = os.path.join(folder_path, 'spectrum.jpg')
    plt.savefig(image_path, dpi=300, bbox_inches='tight')

    # 關閉圖像以釋放資源
    plt.close()

def save_excel(folder_path,wavelength,TR,TG1,TG2,TB):
    from openpyxl.chart import (
        Reference,
        Series,
        LineChart,
    )

    # 創建一個新的Excel工作簿
    workbook = openpyxl.Workbook()

    # 獲取活躍的工作表
    worksheet = workbook.active

    # 將數據寫入工作表
    worksheet['A1'] = 'Wavelength'
    worksheet['B1'] = 'TR'
    worksheet['C1'] = 'TG1'
    worksheet['D1'] = 'TG2'
    worksheet['E1'] = 'TB'

    # 遍歷數據並寫入工作表
    for row, (w, tr, tg1, tg2, tb) in enumerate(zip(wavelength, TR, TG1, TG2, TB), start=2):
        worksheet.cell(row=row, column=1, value=w)
        worksheet.cell(row=row, column=2, value=tr)
        worksheet.cell(row=row, column=3, value=tg1)
        worksheet.cell(row=row, column=4, value=tg2)
        worksheet.cell(row=row, column=5, value=tb)

    # 創建一個圖表對象
    chart = LineChart()

    # 設置圖表標題
    chart.title = "Spectrum"

    # 設置圖表x軸標籤
    chart.x_axis.title = "Wavelength (nm)"

    # 設置圖表y軸標籤
    chart.y_axis.title = "Intensity (a.u)"

    # 添加數據序列
    xvalues = Reference(worksheet, min_col=1, min_row=2, max_row=len(wavelength)+1)
    values1 = Reference(worksheet, min_col=2, min_row=1, max_row=len(TR)+1)
    values2 = Reference(worksheet, min_col=3, min_row=1, max_row=len(TG1)+1)
    values3 = Reference(worksheet, min_col=4, min_row=1, max_row=len(TG2)+1)
    values4 = Reference(worksheet, min_col=5, min_row=1, max_row=len(TB)+1)
    
    series1 = Series(values1, title_from_data=True)
    series2 = Series(values2, title_from_data=True)
    series3 = Series(values3, title_from_data=True)
    series4 = Series(values4, title_from_data=True)

    chart.append(series1)
    chart.append(series2)
    chart.append(series3)
    chart.append(series4)
    #set x-axis
    chart.set_categories(xvalues)
    chart.x_axis.number_format = '***'
    # 將圖表添加到工作表
    worksheet.add_chart(chart, "G2")
    # 保存工作簿
    workbook.save(f'{folder_path}/spectrum_data.xlsx')
    print(f'Done! {folder_path}')

if __name__ == '__main__':
    folder_name = 'RGB'
    for i in range(1,82):
        filename = f'{i}.png'
        #Wiener_deconv(TestImage = folder_name, filename = filename, SNR_F=300.0)
        #RL_deconv(TestImage = folder_name, filename = filename, max_iter_RL = 30)
        #BRL_deconv(TestImage = folder_name, filename = filename, max_iter_RL = 30, rk = 5, sigma_r = 10.0, lamb_da = 0.0005)
        #TVL1_deconv(TestImage = folder_name,filename = filename, max_iter = 2000, lamb_da = 0.00025)
        #TVL2_deconv(TestImage = folder_name,filename = filename, max_iter = 2000, lamb_da = 0.00025)
        #TVpoisson_deconv(TestImage = folder_name,filename = filename, max_iter = 2000, lamb_da = 0.00025)
        path = f'../data/{folder_name}/result'
    folders = os.listdir(path)
    for folder in folders:
        folder_path = os.path.join(path, folder)
        RGB_generate(folder_path, img_list = [0,22,57])
        TR,TG1,TG2,TB = get_spectrum(folder_path)
        wavelength = np.linspace(470,900,150)
        wavelength = wavelength[0:81]
        plot_spectrum(folder_path,wavelength, TR, TG1, TG2, TB)
        save_excel(folder_path,wavelength,TR,TG1,TG2,TB)